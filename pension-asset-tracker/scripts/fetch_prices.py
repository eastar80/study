# -*- coding: utf-8 -*-
"""
계좌설정에 등록된 ETF들의 최근 종가(한국거래소 공개 데이터)를 받아와
현황_리밸런싱 시트의 '현재가' 컬럼을 자동으로 채운다.

사용법:
    python3 fetch_prices.py [워크북경로]

주의:
    - KRX는 실시간 시세를 무료로 공개하지 않으므로, 최근 영업일 '종가' 기준입니다.
      분기 리밸런싱 목적에는 충분히 정확하지만 장중 실시간가는 아닙니다.
    - 로그인/증권사 계정 연동이 필요 없습니다 (공개 데이터만 사용).
"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from pykrx import stock

from common import find_headers, find_data_rows, normalize_ticker, latest_etf_ohlcv

DEFAULT_FILE = Path(__file__).resolve().parent.parent / "연금자산관리_템플릿.xlsx"
HEADER_ROW = 4


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FILE
    if not path.exists():
        print(f"파일을 찾을 수 없습니다: {path}")
        sys.exit(1)

    wb = load_workbook(path)
    ws_acct = wb["계좌설정"]
    ws_cur = wb["현황_리밸런싱"]

    acct_headers = find_headers(ws_acct, HEADER_ROW)
    cur_headers = find_headers(ws_cur, HEADER_ROW)
    acct_rows = find_data_rows(ws_acct, HEADER_ROW, acct_headers["계좌ID"])
    cur_rows = find_data_rows(ws_cur, HEADER_ROW, cur_headers["계좌ID"])

    if len(acct_rows) != len(cur_rows):
        print("경고: 계좌설정과 현황_리밸런싱의 데이터 행 수가 다릅니다. "
              "두 시트가 같은 순서로 맞춰져 있는지 확인하세요. (README 참고)")

    print("최근 영업일 ETF 종가를 조회합니다 (한국거래소 공개 데이터)...")
    date_str, price_df = latest_etf_ohlcv(stock)
    print(f"기준일: {date_str}")

    updated, skipped = [], []
    for acct_row, cur_row in zip(acct_rows, cur_rows):
        ticker = normalize_ticker(ws_acct.cell(row=acct_row, column=acct_headers["티커"]).value)
        name = ws_acct.cell(row=acct_row, column=acct_headers["ETF명"]).value
        if ticker in price_df.index:
            price = int(price_df.loc[ticker, "종가"])
            cell = ws_cur.cell(row=cur_row, column=cur_headers["현재가"], value=price)
            cell.number_format = '#,##0'
            updated.append((ticker, name, price))
        else:
            skipped.append(f"{ticker}({name})")

    wb.save(path)

    print(f"\n{len(updated)}개 종목 현재가 갱신 완료 → {path.name}")
    for ticker, name, price in updated:
        print(f"  {ticker} {name}: {price:,}원")
    if skipped:
        print(f"\n조회 실패(직접 확인 필요): {', '.join(skipped)}")
        print("→ 티커가 올바른지, 최근 상장폐지/코드변경된 종목은 아닌지 확인하세요.")


if __name__ == "__main__":
    main()
