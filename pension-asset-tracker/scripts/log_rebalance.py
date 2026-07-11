# -*- coding: utf-8 -*-
"""
현황_리밸런싱에서 매수/매도가 필요한 종목을 리밸런싱_히스토리 시트에
오늘 날짜(또는 지정한 날짜)로 초안 기록한다.

사용법:
    python3 log_rebalance.py [워크북경로] [--date YYYY-MM-DD]

동작:
    - 현황_리밸런싱에 이미 입력된 보유수량·현재가를 읽어 계좌설정의 목표비중과
      비교, 액션(매수/매도)이 발생하는 종목만 리밸런싱_히스토리 맨 아래에 추가합니다.
    - 체결단가는 조회 시점의 현재가로 '추정' 기록되므로, 실제 주문 체결 후
      리밸런싱_히스토리에서 체결단가(및 수량이 다르면 수량도)를 실제 값으로 수정하세요.
    - 네트워크가 필요 없습니다 (시트에 이미 입력된 값만 사용).
"""
import sys
import argparse
import datetime
from pathlib import Path

from openpyxl import load_workbook

from common import find_headers, find_data_rows, normalize_ticker

DEFAULT_FILE = Path(__file__).resolve().parent.parent / "연금자산관리_템플릿.xlsx"
HEADER_ROW = 4


def compute_trades(ws_acct, ws_cur):
    acct_headers = find_headers(ws_acct, HEADER_ROW)
    cur_headers = find_headers(ws_cur, HEADER_ROW)
    acct_rows = find_data_rows(ws_acct, HEADER_ROW, acct_headers["계좌ID"])
    cur_rows = find_data_rows(ws_cur, HEADER_ROW, cur_headers["계좌ID"])

    positions = []
    for acct_row, cur_row in zip(acct_rows, cur_rows):
        acct_id = ws_acct.cell(row=acct_row, column=acct_headers["계좌ID"]).value
        ticker = normalize_ticker(ws_acct.cell(row=acct_row, column=acct_headers["티커"]).value)
        name = ws_acct.cell(row=acct_row, column=acct_headers["ETF명"]).value
        target_w = ws_acct.cell(row=acct_row, column=acct_headers["목표비중"]).value or 0
        qty = ws_cur.cell(row=cur_row, column=cur_headers["보유수량"]).value or 0
        price = ws_cur.cell(row=cur_row, column=cur_headers["현재가"]).value or 0
        positions.append(dict(acct_id=acct_id, ticker=ticker, name=name,
                               target_w=target_w, qty=qty, price=price, value=qty * price))

    account_totals = {}
    for pos in positions:
        account_totals[pos["acct_id"]] = account_totals.get(pos["acct_id"], 0) + pos["value"]

    trades = []
    for pos in positions:
        total = account_totals.get(pos["acct_id"], 0)
        if total == 0 or pos["price"] == 0:
            continue
        target_value = total * pos["target_w"]
        adjust = target_value - pos["value"]
        trade_qty = round(adjust / pos["price"])
        if trade_qty == 0:
            continue
        trades.append(dict(
            acct_id=pos["acct_id"], ticker=pos["ticker"], name=pos["name"],
            action="매수" if trade_qty > 0 else "매도",
            qty=abs(trade_qty), price=pos["price"],
        ))
    return trades


def append_history(ws_hist, trades, date_str):
    headers = find_headers(ws_hist, HEADER_ROW)
    date_col = headers["날짜"]
    r = HEADER_ROW + 1
    while ws_hist.cell(row=r, column=date_col).value not in (None, ""):
        r += 1

    note = "자동생성(log_rebalance.py) - 체결 후 실제 체결단가로 수정하세요"
    for trade in trades:
        ws_hist.cell(row=r, column=headers["날짜"], value=date_str)
        ws_hist.cell(row=r, column=headers["계좌ID"], value=trade["acct_id"])
        ws_hist.cell(row=r, column=headers["티커"], value=trade["ticker"])
        ws_hist.cell(row=r, column=headers["액션"], value=trade["action"])
        ws_hist.cell(row=r, column=headers["수량"], value=trade["qty"])
        price_cell = ws_hist.cell(row=r, column=headers["체결단가"], value=trade["price"])
        price_cell.number_format = '#,##0'
        ws_hist.cell(row=r, column=headers["메모"], value=note)
        r += 1


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("file", nargs="?", default=str(DEFAULT_FILE))
    parser.add_argument("--date", default=None, help="YYYY-MM-DD (기본값: 오늘)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"파일을 찾을 수 없습니다: {path}")
        sys.exit(1)

    date_str = args.date or datetime.date.today().strftime("%Y-%m-%d")

    wb = load_workbook(path)
    ws_acct = wb["계좌설정"]
    ws_cur = wb["현황_리밸런싱"]
    ws_hist = wb["리밸런싱_히스토리"]

    trades = compute_trades(ws_acct, ws_cur)
    if not trades:
        print("매수/매도가 필요한 종목이 없습니다. "
              "현황_리밸런싱에 보유수량·현재가가 입력되어 있는지 확인하세요.")
        return

    append_history(ws_hist, trades, date_str)
    wb.save(path)

    print(f"{len(trades)}건의 매수/매도 초안을 리밸런싱_히스토리에 기록했습니다 (날짜: {date_str}).")
    for t in trades:
        print(f"  [{t['acct_id']}] {t['action']} {t['name']}({t['ticker']}) "
              f"{t['qty']}주 @ 약 {t['price']:,}원(추정)")
    print("\n실제 체결 후 리밸런싱_히스토리에서 체결단가(및 필요시 수량)를 실제 값으로 수정하세요.")


if __name__ == "__main__":
    main()
