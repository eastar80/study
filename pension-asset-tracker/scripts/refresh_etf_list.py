# -*- coding: utf-8 -*-
"""
ETF전체목록 시트를 한국거래소(KRX)에 상장된 전체 ETF 목록(티커·이름·기초지수·최근종가)으로 갱신한다.
종목을 새로 추가할 때 이름으로 찾아 티커를 계좌설정에 복사해 쓰기 위한 참고용 시트입니다.

사용법:
    python3 refresh_etf_list.py [워크북경로]

주의:
    - 로그인/증권사 계정 연동이 필요 없습니다 (공개 데이터만 사용).
    - 상장 ETF가 900개 안팎이라 이름 조회에 다소 시간이 걸릴 수 있습니다(약 1~2분).
"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from pykrx import stock

from common import latest_etf_ohlcv

DEFAULT_FILE = Path(__file__).resolve().parent.parent / "연금자산관리_템플릿.xlsx"
SHEET_NAME = "ETF전체목록"
HEADER_ROW = 5


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FILE
    if not path.exists():
        print(f"파일을 찾을 수 없습니다: {path}")
        sys.exit(1)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]

    print("상장 ETF 목록 및 종가를 조회합니다 (한국거래소 공개 데이터)...")
    date_str, df = latest_etf_ohlcv(stock)
    print(f"기준일: {date_str}, 총 {len(df)}개 종목. 종목명 조회 중...")

    names = {ticker: stock.get_etf_ticker_name(ticker) for ticker in df.index}
    tickers = sorted(df.index, key=lambda t: names[t])

    # 기존 데이터 삭제
    max_clear_row = max(HEADER_ROW + len(tickers) + 10, HEADER_ROW + 1000)
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=max_clear_row, max_col=4):
        for cell in row:
            cell.value = None

    r = HEADER_ROW + 1
    for ticker in tickers:
        base_index = df.loc[ticker, "기초지수"] if "기초지수" in df.columns else None
        close = int(df.loc[ticker, "종가"])
        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=names[ticker])
        ws.cell(row=r, column=3, value=base_index)
        price_cell = ws.cell(row=r, column=4, value=close)
        price_cell.number_format = '#,##0'
        r += 1

    ws["A3"] = f"기준일: {date_str} (총 {len(tickers)}개 종목) — refresh_etf_list.py 로 갱신됨"
    ws.auto_filter.ref = f"A{HEADER_ROW}:D{r - 1}"

    wb.save(path)
    print(f"ETF전체목록 시트에 {len(tickers)}개 종목을 기록했습니다 → {path.name}")


if __name__ == "__main__":
    main()
