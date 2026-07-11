# -*- coding: utf-8 -*-
"""
연금 자산(IRP/연금저축) 4계좌 ETF 리밸런싱 관리 템플릿 생성 스크립트.
실행: python3 generate_template.py
결과: 연금자산관리_템플릿.xlsx 생성/갱신
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# ---------- 공통 스타일 ----------
NAVY = "1F3864"
LIGHT_BLUE = "DDEBF7"
GREEN = "C6E0B4"
RED = "F8CBAD"
GREY = "F2F2F2"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, size=16, color=NAVY)
SUB_FONT = Font(bold=True, size=11, color=NAVY)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

PCT = '0.0%'
KRW = '#,##0"원"'
KRW0 = '#,##0'

def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_title(ws, cell_ref, text, size=16):
    ws[cell_ref] = text
    ws[cell_ref].font = Font(bold=True, size=size, color=NAVY)

def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER

wb = Workbook()

# =====================================================================
# 0. 사용법
# =====================================================================
ws0 = wb.active
ws0.title = "사용법"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 4
ws0.column_dimensions["B"].width = 100

style_title(ws0, "B2", "연금 자산(IRP·연금저축) 리밸런싱 관리 템플릿", 18)
ws0["B3"] = "4개 계좌 × 다수 ETF, 분기별 리밸런싱을 관리하기 위한 엑셀 템플릿입니다."
ws0["B3"].font = Font(italic=True, color="666666")

guide = [
    ("① 계좌설정", "4개 계좌(IRP/연금저축)와 각 계좌가 보유할 ETF 목록·목표비중(%)을 등록합니다. "
                 "계좌당 목표비중 합계는 100%가 되어야 합니다. (예시 데이터가 들어있으니 실제 보유 종목으로 교체하세요)"),
    ("② 현황_리밸런싱", "분기마다 계좌설정과 '동일한 순서'로 보유수량·현재가만 입력하면, 평가금액/현재비중/목표비중과의 "
                    "괴리, 계좌별 매수·매도 필요 수량이 자동 계산됩니다. 이 결과를 보고 실제 주문을 넣으면 됩니다."),
    ("③ 리밸런싱_히스토리", "실제로 체결한 매수/매도 내역을 한 줄씩 기록하는 로그입니다. 계좌별칭·ETF명은 계좌설정을 "
                        "참조해 자동으로 채워집니다. 과거에 무엇을 언제 왜 했는지 전부 남습니다."),
    ("④ 시계열_스냅샷", "분기가 끝날 때마다(리밸런싱 직후) 계좌별 평가금액 합계를 한 줄 추가하세요. "
                     "총자산 추이와 계좌별 비중 변화를 시계열로 볼 수 있습니다."),
    ("⑤ 대시보드", "현재 총자산, 계좌별 요약, 목표 대비 괴리가 큰 종목을 한눈에 보여주는 요약 화면입니다."),
    ("⑥ 차트", "총자산 추이, 계좌별 자산 추이, 목표비중 대비 현재비중, 계좌별 자산배분 파이차트를 모아둔 시트입니다."),
]
r = 5
for title, desc in guide:
    ws0.cell(row=r, column=2, value=title).font = Font(bold=True, size=12, color=NAVY)
    r += 1
    c = ws0.cell(row=r, column=2, value=desc)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.row_dimensions[r].height = 32
    r += 2

ws0.cell(row=r, column=2, value="권장 루틴 (분기 1회, 약 10분)").font = Font(bold=True, size=12, color=NAVY)
r += 1
for step in [
    "1) 증권사 앱에서 4계좌 각각의 보유수량·현재가를 확인",
    "2) 현황_리밸런싱 시트의 보유수량/현재가 입력 → 매수·매도 수량 자동 산출 확인",
    "3) 산출된 대로 각 계좌에서 실제 주문 실행",
    "4) 체결 내역을 리밸런싱_히스토리에 한 줄씩 기록",
    "5) 시계열_스냅샷에 이번 분기 계좌별 합계금액 한 줄 추가",
]:
    ws0.cell(row=r, column=2, value=step)
    r += 1

# =====================================================================
# 1. 계좌설정
# =====================================================================
ws1 = wb.create_sheet("계좌설정")
ws1.sheet_view.showGridLines = False
widths = [10, 12, 16, 16, 10, 26, 12]
for i, w in enumerate(widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

style_title(ws1, "A1", "계좌 & 목표 포트폴리오 설정")
ws1["A2"] = "※ 예시 데이터입니다. 실제 보유 계좌/ETF/목표비중으로 교체하세요. 계좌당 목표비중 합계는 100%가 되어야 합니다."
ws1["A2"].font = Font(italic=True, size=9, color="C00000")

headers1 = ["계좌ID", "계좌구분", "계좌별칭", "증권사", "티커", "ETF명", "목표비중"]
HEADER_ROW1 = 4
for i, h in enumerate(headers1, start=1):
    ws1.cell(row=HEADER_ROW1, column=i, value=h)
style_header(ws1, HEADER_ROW1, 1, len(headers1))

sample_accounts = [
    ("A1", "IRP", "IRP-1(미래에셋)", "미래에셋증권", [
        ("379800", "KODEX 미국S&P500", 0.40),
        ("381170", "TIGER 미국테크TOP10", 0.20),
        ("360750", "TIGER 미국나스닥100", 0.20),
        ("305080", "TIGER 미국채10년선물", 0.20),
    ]),
    ("A2", "연금저축", "연금저축-1(미래에셋)", "미래에셋증권", [
        ("069500", "KODEX 200", 0.40),
        ("148070", "KOSEF 국고채10년", 0.30),
        ("132030", "KODEX 골드선물(H)", 0.15),
        ("310970", "KODEX 미국S&P500TR", 0.15),
    ]),
    ("B1", "IRP", "IRP-2(한국투자)", "한국투자증권", [
        ("360750", "TIGER 미국나스닥100", 0.30),
        ("379800", "KODEX 미국S&P500", 0.30),
        ("114260", "KODEX 국고채3년", 0.25),
        ("132030", "KODEX 골드선물(H)", 0.15),
    ]),
    ("B2", "연금저축", "연금저축-2(한국투자)", "한국투자증권", [
        ("069500", "KODEX 200", 0.35),
        ("360750", "TIGER 미국나스닥100", 0.25),
        ("148070", "KOSEF 국고채10년", 0.25),
        ("130680", "TIGER 골드선물(H)", 0.15),
    ]),
]

DATA_START1 = HEADER_ROW1 + 1  # 5
row = DATA_START1
ACCOUNT_ROW_SPAN = 4  # 계좌당 ETF 4종 고정 (필요시 행 추가 가능)
account_alias_rows = []  # 각 계좌 첫 행 번호 기록 (시계열 헤더 링크용)
for acct_id, acct_type, alias, broker, etfs in sample_accounts:
    account_alias_rows.append(row)
    for j, (ticker, name, weight) in enumerate(etfs):
        ws1.cell(row=row, column=1, value=acct_id)
        ws1.cell(row=row, column=2, value=acct_type)
        ws1.cell(row=row, column=3, value=alias)
        ws1.cell(row=row, column=4, value=broker)
        ws1.cell(row=row, column=5, value=ticker)
        ws1.cell(row=row, column=6, value=name)
        wcell = ws1.cell(row=row, column=7, value=weight)
        wcell.number_format = PCT
        row += 1
DATA_END1 = row - 1  # 20

# 계좌별 목표비중 합계 = 100% 검증행
CHECK_ROW1 = DATA_END1 + 2
ws1.cell(row=CHECK_ROW1, column=3, value="계좌별 목표비중 합계 검증 (100%가 아니면 빨간색)")
ws1.cell(row=CHECK_ROW1, column=3).font = Font(bold=True, size=9, color="666666")
chk_col = 8
ws1.cell(row=CHECK_ROW1 - 1, column=chk_col, value="계좌ID")
ws1.cell(row=CHECK_ROW1 - 1, column=chk_col + 1, value="목표비중합계")
style_header(ws1, CHECK_ROW1 - 1, chk_col, chk_col + 1)
for i, (acct_id, *_rest) in enumerate(sample_accounts):
    rr = CHECK_ROW1 + i
    ws1.cell(row=rr, column=chk_col, value=acct_id)
    f = ws1.cell(row=rr, column=chk_col + 1,
                 value=f"=SUMIF($A${DATA_START1}:$A${DATA_END1},{get_column_letter(chk_col)}{rr},"
                       f"$G${DATA_START1}:$G${DATA_END1})")
    f.number_format = PCT

box(ws1, HEADER_ROW1, 1, DATA_END1, 7)
ws1.freeze_panes = f"A{DATA_START1}"

# 조건부서식: 목표비중 합계가 100%가 아니면 빨간 글씨
ws1.conditional_formatting.add(
    f"{get_column_letter(chk_col+1)}{CHECK_ROW1}:{get_column_letter(chk_col+1)}{CHECK_ROW1+3}",
    CellIsRule(operator="notEqual", formula=["1"], font=Font(color="C00000", bold=True))
)

# 데이터 검증(드롭다운)
dv_type = DataValidation(type="list", formula1='"IRP,연금저축,ISA,일반계좌"', allow_blank=True)
ws1.add_data_validation(dv_type)
dv_type.add(f"B{DATA_START1}:B{DATA_END1+20}")

# =====================================================================
# 2. 현황_리밸런싱
# =====================================================================
ws2 = wb.create_sheet("현황_리밸런싱")
ws2.sheet_view.showGridLines = False
widths2 = [9, 16, 10, 24, 12, 12, 14, 14, 12, 12, 10, 14, 14, 12, 8, 10]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

style_title(ws2, "A1", "분기 리밸런싱 계산기")
ws2["A2"] = "※ 계좌설정과 '동일한 순서'로 보유수량·현재가만 입력하세요. 나머지는 전부 자동 계산됩니다."
ws2["A2"].font = Font(italic=True, size=9, color="C00000")

headers2 = ["계좌ID", "계좌별칭", "티커", "ETF명", "보유수량", "현재가", "평가금액",
            "계좌합계금액", "현재비중", "목표비중", "비중차이", "목표금액", "조정금액(+매수/-매도)",
            "매매수량", "액션", "비중차이(절대값)"]
HEADER_ROW2 = 4
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=HEADER_ROW2, column=i, value=h)
style_header(ws2, HEADER_ROW2, 1, len(headers2))

DATA_START2 = HEADER_ROW2 + 1
n_rows = DATA_END1 - DATA_START1 + 1
DATA_END2 = DATA_START2 + n_rows - 1

for offset in range(n_rows):
    r2 = DATA_START2 + offset
    r1 = DATA_START1 + offset
    ws2.cell(row=r2, column=1, value=f"=계좌설정!A{r1}")
    ws2.cell(row=r2, column=2, value=f"=계좌설정!C{r1}")
    ws2.cell(row=r2, column=3, value=f"=계좌설정!E{r1}")
    ws2.cell(row=r2, column=4, value=f"=계좌설정!F{r1}")
    # E,F 보유수량/현재가는 입력칸(빈칸) + 노란 배경
    ws2.cell(row=r2, column=5).fill = PatternFill("solid", fgColor="FFF2CC")
    ws2.cell(row=r2, column=6).fill = PatternFill("solid", fgColor="FFF2CC")
    ws2.cell(row=r2, column=6).number_format = KRW0
    g = ws2.cell(row=r2, column=7, value=f"=E{r2}*F{r2}")
    g.number_format = KRW
    h = ws2.cell(row=r2, column=8,
                 value=f"=SUMIF($A${DATA_START2}:$A${DATA_END2},A{r2},$G${DATA_START2}:$G${DATA_END2})")
    h.number_format = KRW
    i_ = ws2.cell(row=r2, column=9, value=f"=IF(H{r2}=0,0,G{r2}/H{r2})")
    i_.number_format = PCT
    j = ws2.cell(row=r2, column=10, value=f"=계좌설정!G{r1}")
    j.number_format = PCT
    k = ws2.cell(row=r2, column=11, value=f"=I{r2}-J{r2}")
    k.number_format = PCT
    l = ws2.cell(row=r2, column=12, value=f"=H{r2}*J{r2}")
    l.number_format = KRW
    m = ws2.cell(row=r2, column=13, value=f"=L{r2}-G{r2}")
    m.number_format = KRW
    n = ws2.cell(row=r2, column=14, value=f"=IF(F{r2}=0,0,ROUND(M{r2}/F{r2},0))")
    o = ws2.cell(row=r2, column=15, value=f'=IF(N{r2}>0,"매수",IF(N{r2}<0,"매도","유지"))')
    o.alignment = CENTER
    p = ws2.cell(row=r2, column=16, value=f"=ABS(K{r2})")
    p.number_format = PCT

# 합계 행
TOTAL_ROW2 = DATA_END2 + 1
ws2.cell(row=TOTAL_ROW2, column=4, value="총계").font = Font(bold=True)
tot_g = ws2.cell(row=TOTAL_ROW2, column=7, value=f"=SUM(G{DATA_START2}:G{DATA_END2})")
tot_g.number_format = KRW
tot_g.font = Font(bold=True)
ws2.cell(row=TOTAL_ROW2, column=7).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
for col in (4,):
    ws2.cell(row=TOTAL_ROW2, column=col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

box(ws2, HEADER_ROW2, 1, TOTAL_ROW2, 16)
ws2.freeze_panes = f"E{DATA_START2}"
ws2.column_dimensions["P"].hidden = True

# 조건부서식: 비중차이 절대값 5%p 초과시 강조, 액션 매수=초록/매도=빨강
ws2.conditional_formatting.add(
    f"K{DATA_START2}:K{DATA_END2}",
    CellIsRule(operator="greaterThan", formula=["0.05"], fill=PatternFill("solid", fgColor=RED))
)
ws2.conditional_formatting.add(
    f"K{DATA_START2}:K{DATA_END2}",
    CellIsRule(operator="lessThan", formula=["-0.05"], fill=PatternFill("solid", fgColor=RED))
)
ws2.conditional_formatting.add(
    f"O{DATA_START2}:O{DATA_END2}",
    CellIsRule(operator="equal", formula=['"매수"'], fill=PatternFill("solid", fgColor=GREEN))
)
ws2.conditional_formatting.add(
    f"O{DATA_START2}:O{DATA_END2}",
    CellIsRule(operator="equal", formula=['"매도"'], fill=PatternFill("solid", fgColor=RED))
)

# =====================================================================
# 3. 리밸런싱_히스토리
# =====================================================================
ws3 = wb.create_sheet("리밸런싱_히스토리")
ws3.sheet_view.showGridLines = False
widths3 = [12, 9, 10, 16, 10, 24, 8, 10, 12, 14, 24]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

style_title(ws3, "A1", "리밸런싱 실행 히스토리 (체결 로그)")
ws3["A2"] = "실제로 매수/매도를 체결한 후 한 줄씩 기록하세요. 계좌별칭·ETF명은 계좌설정을 참조해 자동 표시됩니다."
ws3["A2"].font = Font(italic=True, size=9, color="666666")

headers3 = ["날짜", "분기", "계좌ID", "계좌별칭", "티커", "ETF명", "액션", "수량", "체결단가", "거래금액", "메모"]
HEADER_ROW3 = 4
for i, h in enumerate(headers3, start=1):
    ws3.cell(row=HEADER_ROW3, column=i, value=h)
style_header(ws3, HEADER_ROW3, 1, len(headers3))

DATA_START3 = HEADER_ROW3 + 1
N_LOG_ROWS = 60
DATA_END3 = DATA_START3 + N_LOG_ROWS - 1

for r3 in range(DATA_START3, DATA_END3 + 1):
    ws3.cell(row=r3, column=2, value=f'=IF(A{r3}="","",YEAR(A{r3})&"Q"&ROUNDUP(MONTH(A{r3})/3,0))')
    ws3.cell(row=r3, column=4,
             value=f'=IF(C{r3}="","",VLOOKUP(C{r3},계좌설정!$A:$C,3,FALSE))')
    ws3.cell(row=r3, column=6,
             value=f'=IF(E{r3}="","",VLOOKUP(E{r3},계좌설정!$E:$F,2,FALSE))')
    amt = ws3.cell(row=r3, column=10, value=f'=IF(OR(H{r3}="",I{r3}=""),"",H{r3}*I{r3})')
    amt.number_format = KRW
    ws3.cell(row=r3, column=1).number_format = "yyyy-mm-dd"
    ws3.cell(row=r3, column=9).number_format = KRW0

# 예시 한 줄
ws3.cell(row=DATA_START3, column=1, value="2026-07-04")
ws3.cell(row=DATA_START3, column=3, value="A1")
ws3.cell(row=DATA_START3, column=5, value="379800")
ws3.cell(row=DATA_START3, column=7, value="매수")
ws3.cell(row=DATA_START3, column=8, value=5)
ws3.cell(row=DATA_START3, column=9, value=18500)
ws3.cell(row=DATA_START3, column=11, value="분기 리밸런싱 (예시 행 - 삭제 후 사용)")

box(ws3, HEADER_ROW3, 1, DATA_END3, 11)
ws3.freeze_panes = f"A{DATA_START3}"

dv_acct = DataValidation(type="list", formula1='"A1,A2,B1,B2"', allow_blank=True)
ws3.add_data_validation(dv_acct)
dv_acct.add(f"C{DATA_START3}:C{DATA_END3}")
dv_action = DataValidation(type="list", formula1='"매수,매도"', allow_blank=True)
ws3.add_data_validation(dv_action)
dv_action.add(f"G{DATA_START3}:G{DATA_END3}")

# =====================================================================
# 4. 시계열_스냅샷
# =====================================================================
ws4 = wb.create_sheet("시계열_스냅샷")
ws4.sheet_view.showGridLines = False
widths4 = [12, 16, 20, 16, 16, 14, 14, 12]
for i, w in enumerate(widths4, start=1):
    ws4.column_dimensions[get_column_letter(i)].width = w

style_title(ws4, "A1", "분기별 총자산 시계열 스냅샷")
ws4["A2"] = "매 분기 리밸런싱 직후, 계좌별 평가금액 합계를 한 줄씩 '값으로' 입력하세요 (현황_리밸런싱의 계좌합계금액을 복사)."
ws4["A2"].font = Font(italic=True, size=9, color="666666")

HEADER_ROW4 = 4
ws4.cell(row=HEADER_ROW4, column=1, value="날짜")
for i, r1 in enumerate(account_alias_rows):
    ws4.cell(row=HEADER_ROW4, column=2 + i, value=f"=계좌설정!C{r1}")
ws4.cell(row=HEADER_ROW4, column=6, value="총자산")
ws4.cell(row=HEADER_ROW4, column=7, value="전기대비증감액")
ws4.cell(row=HEADER_ROW4, column=8, value="전기대비증감률")
style_header(ws4, HEADER_ROW4, 1, 8)

DATA_START4 = HEADER_ROW4 + 1
N_SNAP_ROWS = 20
DATA_END4 = DATA_START4 + N_SNAP_ROWS - 1

for r4 in range(DATA_START4, DATA_END4 + 1):
    ws4.cell(row=r4, column=1).number_format = "yyyy-mm-dd"
    tot = ws4.cell(row=r4, column=6, value=f"=IF(SUM(B{r4}:E{r4})=0,\"\",SUM(B{r4}:E{r4}))")
    tot.number_format = KRW
    if r4 == DATA_START4:
        ws4.cell(row=r4, column=7, value="")
        ws4.cell(row=r4, column=8, value="")
    else:
        diff = ws4.cell(row=r4, column=7,
                         value=f'=IF(OR(F{r4}="",F{r4-1}=""),"",F{r4}-F{r4-1})')
        diff.number_format = KRW
        pct = ws4.cell(row=r4, column=8,
                        value=f'=IF(OR(F{r4}="",F{r4-1}="",F{r4-1}=0),"",F{r4}/F{r4-1}-1)')
        pct.number_format = PCT
    for col in range(2, 6):
        ws4.cell(row=r4, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

box(ws4, HEADER_ROW4, 1, DATA_END4, 8)
ws4.freeze_panes = f"B{DATA_START4}"

# =====================================================================
# 5. 대시보드
# =====================================================================
ws5 = wb.create_sheet("대시보드")
ws5.sheet_view.showGridLines = False
wb.move_sheet("대시보드", offset=-4)  # 사용법 다음, 계좌설정 이전으로 이동

for col, w in zip("ABCDEFG", [4, 22, 18, 18, 14, 14, 4]):
    ws5.column_dimensions[col].width = w

style_title(ws5, "B2", "연금 자산 대시보드", 18)
ws5["B3"] = "=\"기준일: \"&TEXT(TODAY(),\"yyyy-mm-dd\")"
ws5["B3"].font = Font(italic=True, color="666666")

ws5["B5"] = "총자산"
ws5["B5"].font = SUB_FONT
ws5["B6"] = f"=현황_리밸런싱!G{TOTAL_ROW2}"
ws5["B6"].number_format = KRW
ws5["B6"].font = Font(bold=True, size=20, color=NAVY)

ws5["D5"] = "최근 리밸런싱 실행일"
ws5["D5"].font = SUB_FONT
ws5["D6"] = f"=IFERROR(MAX(리밸런싱_히스토리!A{DATA_START3}:A{DATA_END3}),\"기록 없음\")"
ws5["D6"].number_format = "yyyy-mm-dd"
ws5["D6"].font = Font(bold=True, size=14)

ws5["F5"] = "최근 스냅샷 총자산"
ws5["F5"].font = SUB_FONT
ws5["F6"] = f"=IFERROR(LOOKUP(2,1/(시계열_스냅샷!F{DATA_START4}:F{DATA_END4}<>\"\"),시계열_스냅샷!F{DATA_START4}:F{DATA_END4}),\"데이터 없음\")"
ws5["F6"].number_format = KRW
ws5["F6"].font = Font(bold=True, size=14)

# 계좌별 요약 표
ws5["B9"] = "계좌별 요약"
ws5["B9"].font = SUB_FONT
sum_headers = ["계좌별칭", "계좌구분", "평가금액", "총자산 대비 비중"]
for i, h in enumerate(sum_headers):
    ws5.cell(row=10, column=2 + i, value=h)
style_header(ws5, 10, 2, 5)

for i, r1 in enumerate(account_alias_rows):
    rr = 11 + i
    ws5.cell(row=rr, column=2, value=f"=계좌설정!C{r1}")
    ws5.cell(row=rr, column=3, value=f"=계좌설정!B{r1}")
    val = ws5.cell(row=rr, column=4,
                    value=f"=SUMIF(현황_리밸런싱!$A${DATA_START2}:$A${DATA_END2},계좌설정!A{r1},"
                          f"현황_리밸런싱!$G${DATA_START2}:$G${DATA_END2})")
    val.number_format = KRW
    pct = ws5.cell(row=rr, column=5, value=f"=IF($B$6=0,0,D{rr}/$B$6)")
    pct.number_format = PCT
box(ws5, 10, 2, 10 + len(account_alias_rows), 5)

# 목표 대비 괴리 TOP5
top_row = 10 + len(account_alias_rows) + 3
ws5.cell(row=top_row, column=2, value="목표비중 대비 괴리 TOP5 (리밸런싱 우선순위)").font = SUB_FONT
top_headers = ["계좌별칭", "ETF명", "현재비중", "목표비중", "괴리(%p)"]
for i, h in enumerate(top_headers):
    ws5.cell(row=top_row + 1, column=2 + i, value=h)
style_header(ws5, top_row + 1, 2, 6)
for i in range(5):
    rr = top_row + 2 + i
    abs_range = f"현황_리밸런싱!$P${DATA_START2}:$P${DATA_END2}"
    match_key = f"MATCH(LARGE({abs_range},{i + 1}),{abs_range},0)"
    ws5.cell(row=rr, column=2,
              value=f"=INDEX(현황_리밸런싱!$B${DATA_START2}:$B${DATA_END2},{match_key})")
    ws5.cell(row=rr, column=3,
              value=f"=INDEX(현황_리밸런싱!$D${DATA_START2}:$D${DATA_END2},{match_key})")
    c4 = ws5.cell(row=rr, column=4,
              value=f"=INDEX(현황_리밸런싱!$I${DATA_START2}:$I${DATA_END2},{match_key})")
    c4.number_format = PCT
    c5 = ws5.cell(row=rr, column=5,
              value=f"=INDEX(현황_리밸런싱!$J${DATA_START2}:$J${DATA_END2},{match_key})")
    c5.number_format = PCT
    c6 = ws5.cell(row=rr, column=6,
              value=f"=INDEX(현황_리밸런싱!$K${DATA_START2}:$K${DATA_END2},{match_key})")
    c6.number_format = PCT
box(ws5, top_row + 1, 2, top_row + 6, 6)

# =====================================================================
# 6. 차트
# =====================================================================
ws6 = wb.create_sheet("차트")
ws6.sheet_view.showGridLines = False
style_title(ws6, "A1", "자산 추이 & 배분 차트")

# 6-1. 총자산 추이 (라인차트)
line = LineChart()
line.title = "총자산 추이"
line.style = 10
line.y_axis.title = "총자산(원)"
line.x_axis.title = "날짜"
line.height = 9
line.width = 18
data = Reference(ws4, min_col=6, min_row=HEADER_ROW4, max_row=DATA_END4)
cats = Reference(ws4, min_col=1, min_row=DATA_START4, max_row=DATA_END4)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
ws6.add_chart(line, "A3")

# 6-2. 계좌별 자산 추이 (누적 막대)
bar_stack = BarChart()
bar_stack.type = "col"
bar_stack.grouping = "stacked"
bar_stack.overlap = 100
bar_stack.title = "계좌별 자산 추이"
bar_stack.style = 10
bar_stack.height = 9
bar_stack.width = 18
data2 = Reference(ws4, min_col=2, max_col=5, min_row=HEADER_ROW4, max_row=DATA_END4)
bar_stack.add_data(data2, titles_from_data=True)
bar_stack.set_categories(cats)
ws6.add_chart(bar_stack, "K3")

# 6-3. 목표비중 vs 현재비중 (그룹 막대)
bar_cmp = BarChart()
bar_cmp.type = "col"
bar_cmp.grouping = "clustered"
bar_cmp.title = "종목별 현재비중 vs 목표비중"
bar_cmp.style = 10
bar_cmp.height = 9
bar_cmp.width = 24
cur_data = Reference(ws2, min_col=9, max_col=10, min_row=HEADER_ROW2, max_row=DATA_END2)
cmp_cats = Reference(ws2, min_col=4, min_row=DATA_START2, max_row=DATA_END2)
bar_cmp.add_data(cur_data, titles_from_data=True)
bar_cmp.set_categories(cmp_cats)
ws6.add_chart(bar_cmp, "A21")

# 6-4. 계좌별 자산배분 파이차트
pie = PieChart()
pie.title = "계좌별 자산배분"
pie.height = 9
pie.width = 12
n_acct = len(account_alias_rows)
pie_data = Reference(ws5, min_col=4, min_row=10, max_row=10 + n_acct)
pie_cats = Reference(ws5, min_col=2, min_row=11, max_row=10 + n_acct)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws6.add_chart(pie, "K21")

# =====================================================================
# 시트 순서 & 활성 시트
# =====================================================================
wb.active = 0
out_path = "연금자산관리_템플릿.xlsx"
wb.save(out_path)
print(f"생성 완료: {out_path}")
